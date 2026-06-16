# first do pip install openpyxl
#run in python 3.12.12
import os
import pandas as pd
from functools import reduce
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Folder containing the files
folder = r"Z:\Joe\AnalysisCode_forSharing\ExampleData\FullyAnalyzed\Analysis_withAmyCode_cell1\AlignmentAndChecking\BeforeCorrections\AlignmentOverlap_BestThreshold\Analysis_withAmyCode_cell1_branch1"

# User-defined branch number
branch_number = 1

# User-defined image numbers
image_numbers = [0, 1, 2, 3, 4, 5]

# Generate full file paths
files = [
    os.path.join(folder, f"Image{image}_branch{branch_number}_snapped_bouton_overlap.csv")
    for image in image_numbers
]

output_file = os.path.join(folder, f"Concatenated_branch{branch_number}_snapped_bouton_overlap.csv")

# Read each file, keep only ID and type, and rename type column
dataframes = []
for i, file in enumerate(files):
    df = pd.read_csv(file)

    required_columns = ["label", "type", "bouton_overlap_frac_r2"]
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Column '{col}' not found in file: {file}")

    # Keep only the needed columns
    df = df[["label", "type", "bouton_overlap_frac_r2"]]

    # Rename columns to include image number
    df = df.rename(columns={
        "type": f"Image{i}_type",
        "bouton_overlap_frac_r2": f"Image{i}_bouton_overlap_frac_r2"
    })

    dataframes.append(df)

# Merge all dataframes on ID using outer join
merged_df = reduce(lambda left, right: pd.merge(left, right, on="label", how="outer"), dataframes)

# Sort by ID and fill missing values with blanks
merged_df = merged_df.sort_values(by="label").fillna("")

# Optional: force a clean column order
column_order = ["label"]
for i in image_numbers:
    column_order.extend([
        f"Image{i}_type",
        f"Image{i}_bouton_overlap_frac_r2"
    ])

merged_df = merged_df[column_order]

# Save as Excel
output_file = os.path.join(folder, f"Concatenated_branch{branch_number}_type_only.xlsx")
merged_df.to_excel(output_file, index=False)

# Load workbook for formatting
wb = load_workbook(output_file)
ws = wb.active

# Define fill colors
color_map = {
    "Shaft_SynTd": PatternFill(fill_type="solid", start_color="00FFFF", end_color="00FFFF"),           # cyan
    "Empty_shaft": PatternFill(fill_type="solid", start_color="C0C0C0", end_color="C0C0C0"),          # gray
    "Shaft_SyntdNotScored": PatternFill(fill_type="solid", start_color="00FF00", end_color="00FF00"), # green
    "Empty_spine": PatternFill(fill_type="solid", start_color="C0C0C0", end_color="C0C0C0"),          # gray
    "Spine_SynTd": PatternFill(fill_type="solid", start_color="000080", end_color="000080"),          # navy
    "Spine_SyntdNotScored": PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00"), # yellow
    "Ambiguous": PatternFill(fill_type="solid", start_color="FF00FF", end_color="FF00FF")             # magenta
}

# Fonts
white_font = Font(color="FFFFFF")
blue_font = Font(color="0000FF")

# Get column headers from row 1
headers = {cell.column: cell.value for cell in ws[1]}

# Apply formatting
for row in ws.iter_rows(min_row=2):  # skip header row
    for cell in row:
        header = headers[cell.column]

        # 1. Color code text categories
        if cell.value in color_map:
            cell.fill = color_map[cell.value]

            # White text for dark fills
            if cell.value in ["Spine_SynTd", "Ambiguous"]:
                cell.font = white_font

        # 2. Make numeric bouton_overlap values > 0.5 blue font
        elif header is not None and "bouton_overlap_frac_r2" in str(header):
            try:
                if cell.value is not None and cell.value != "" and float(cell.value) > 0.5:
                    cell.font = blue_font
            except (ValueError, TypeError):
                pass

# Save formatted workbook
wb.save(output_file)

print(f"Concatenated and formatted file saved to: {output_file}")